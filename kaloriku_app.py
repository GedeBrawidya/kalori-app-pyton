# Kaloriku - Aplikasi Kalkulator Kalori Makanan dengan GUI Tkinter
# Nama: kezia Febiola Angriani
# NPM: 230612265

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import matplotlib.pyplot as plt
import os
from datetime import datetime

class AplikasiCaloriku:
    def __init__(self, root):
        self.root = root
        self.root.title("Login - Caloriku")
        self.root.geometry("450x350")
        self.root.configure(bg="#f0f0f0")
        
        self.AKUN_USER = {"kezia": "login"}
        
        self.db_bahan = [
            {"nama": "Nasi Putih", "kategori": "Karbohidrat", "kalori": 130},
            {"nama": "Dada Ayam", "kategori": "Protein", "kalori": 165},
            {"nama": "Bayam", "kategori": "Sayuran", "kalori": 23}
        ]
    
        self.data_resep = []

        self.tampilan_login()

    def tampilan_login(self):
        frame_login = tk.Frame(self.root, bg="white", bd=2, relief="groove")
        frame_login.place(relx=0.5, rely=0.5, anchor="center", width=350, height=280)

        tk.Label(frame_login, text="CALORIKU", font=("Arial", 20, "bold"), bg="white", fg="#2E8B57").pack(pady=(20, 5))
        tk.Label(frame_login, text="Silakan Login Dahulu", font=("Arial", 9), bg="white", fg="grey").pack(pady=(0, 20))

        tk.Label(frame_login, text="Username:", bg="white", font=("Arial", 10)).pack(anchor="w", padx=30)
        self.entry_user = tk.Entry(frame_login, bg="#f9f9f9")
        self.entry_user.pack(fill="x", padx=30, pady=5)

        tk.Label(frame_login, text="Password:", bg="white", font=("Arial", 10)).pack(anchor="w", padx=30)
        self.entry_pass = tk.Entry(frame_login, show="*", bg="#f9f9f9")
        self.entry_pass.pack(fill="x", padx=30, pady=5)
        
        tk.Button(frame_login, text="MASUK", command=self.cek_login, bg="#2E8B57", fg="white", font=("Arial", 10, "bold"), width=15).pack(pady=20)

    def cek_login(self):
        user = self.entry_user.get()
        pwd = self.entry_pass.get()

        if user in self.AKUN_USER and self.AKUN_USER[user] == pwd:
            messagebox.showinfo("Berhasil", f"Selamat datang, {user}!")
            self.root.withdraw() 
            self.buka_menu_utama()
        else:
            messagebox.showerror("Gagal", "Username atau Password salah!")

    def buka_menu_utama(self):
        self.window_menu = tk.Toplevel()
        self.window_menu.title("Menu Utama - Caloriku")
        self.window_menu.geometry("500x450")
        self.window_menu.configure(bg="white")

        menubar = tk.Menu(self.window_menu)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Logout", command=self.logout)
        filemenu.add_command(label="Keluar", command=self.keluar_app)
        menubar.add_cascade(label="File", menu=filemenu)
        
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="Info", command=lambda: messagebox.showinfo("Info", "Aplikasi Caloriku"))
        menubar.add_cascade(label="Bantuan", menu=helpmenu)
        self.window_menu.config(menu=menubar)

        tk.Label(self.window_menu, text="DASHBOARD UTAMA", font=("Arial", 18, "bold"), bg="white", fg="#2E8B57").pack(pady=30)
        tk.Label(self.window_menu, text="Pilih Menu Operasional:", font=("Arial", 10), bg="white").pack(pady=(0, 20))
        
        frame_btn = tk.Frame(self.window_menu, bg="white")
        frame_btn.pack()

        tk.Button(frame_btn, text="📝  1. Input Makanan Baru", command=self.buka_input_bahan, bg="#4CAF50", fg="white", font=("Arial", 11), width=30, height=2).pack(pady=10)
        tk.Button(frame_btn, text="🍳  2. Cek Kalori Makanan", command=self.buka_hitung_kalori, bg="#4CAF50", fg="white", font=("Arial", 11), width=30, height=2).pack(pady=10)
        
        tk.Button(self.window_menu, text="Keluar / Logout", command=self.logout, bg="#D32F2F", fg="white", width=20).pack(pady=40)

    def buka_input_bahan(self):
        self.win_input = tk.Toplevel()
        self.win_input.title("Input Bahan Makanan")
        self.win_input.geometry("450x400")
        self.win_input.configure(bg="#f0f0f0")

        tk.Label(self.win_input, text="Tambah Data Bahan", font=("Arial", 14, "bold"), bg="#f0f0f0", fg="#2E8B57").pack(pady=15)

        frame_form = tk.LabelFrame(self.win_input, text="Form Data", bg="white", padx=15, pady=15)
        frame_form.pack(padx=20, fill="x")

        tk.Label(frame_form, text="Nama Makanan:", bg="white").pack(anchor="w")
        self.ent_nama_baru = tk.Entry(frame_form)
        self.ent_nama_baru.pack(fill="x", pady=(0, 10))

        tk.Label(frame_form, text="Kategori:", bg="white").pack(anchor="w")
        self.cb_kategori_baru = ttk.Combobox(frame_form, values=["Karbohidrat", "Protein", "Lemak", "Sayuran", "Buah"], state="readonly")
        self.cb_kategori_baru.pack(fill="x", pady=(0, 10))

        tk.Label(frame_form, text="Kalori (per 100g):", bg="white").pack(anchor="w")
        self.ent_kalori_baru = tk.Entry(frame_form)
        self.ent_kalori_baru.pack(fill="x", pady=(0, 10))

        tk.Button(frame_form, text="Simpan Data", command=self.simpan_bahan_baru, bg="#2E8B57", fg="white").pack(fill="x", pady=10)

    def buka_hitung_kalori(self):
        self.win_hitung = tk.Toplevel()
        self.win_hitung.title("Kalkulator Resep")
        self.win_hitung.geometry("800x600") # Sedikit diperbesar
        self.win_hitung.configure(bg="white")

        menubar_hitung = tk.Menu(self.win_hitung)
        menu_analisis = tk.Menu(menubar_hitung, tearoff=0)
        menu_analisis.add_command(label="Lihat Grafik Visual (Bar & Pie)", command=self.tampilkan_grafik)
        menubar_hitung.add_cascade(label="Analisis / Grafik", menu=menu_analisis)

        menu_file = tk.Menu(menubar_hitung, tearoff=0)
        menu_file.add_command(label="Export ke Excel", command=self.simpan_laporan_excel)
        menubar_hitung.add_cascade(label="File", menu=menu_file)

        self.win_hitung.config(menu=menubar_hitung)

        self.var_nama_resep = tk.StringVar()
        self.var_gram = tk.DoubleVar(value=100)
        self.var_porsi = tk.IntVar(value=1) 

        frame_kiri = tk.LabelFrame(self.win_hitung, text="Input Resep", bg="white", padx=10, pady=10)
        frame_kiri.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        tk.Label(frame_kiri, text="Nama Resep:", bg="white").pack(anchor="w")
        tk.Entry(frame_kiri, textvariable=self.var_nama_resep).pack(fill="x", pady=5)

        tk.Label(frame_kiri, text="Jumlah Porsi:", bg="white").pack(anchor="w", pady=(10, 0))

        self.sb_porsi = tk.Spinbox(frame_kiri, from_=1, to=100, textvariable=self.var_porsi, command=self.update_total_kalori)
        self.sb_porsi.pack(fill="x", pady=5)
        self.sb_porsi.bind("<KeyRelease>", lambda event: self.update_total_kalori())

        tk.Label(frame_kiri, text="Pilih Bahan:", bg="white").pack(anchor="w", pady=(10, 0))
        list_nama = []
        for b in self.db_bahan:
            list_nama.append(b['nama'])
        self.cb_pilih_bahan = ttk.Combobox(frame_kiri, values=list_nama, state="readonly")
        self.cb_pilih_bahan.pack(fill="x", pady=5)

        tk.Label(frame_kiri, text="Berat (gram):", bg="white").pack(anchor="w", pady=(10, 0))
        tk.Scale(frame_kiri, from_=1, to=500, orient="horizontal", variable=self.var_gram, bg="white").pack(fill="x")
        
        tk.Button(frame_kiri, text="+ Tambah ke Tabel", command=self.tambah_ke_resep, bg="#2E8B57", fg="white", height=2).pack(pady=20, fill="x")

        frame_kanan = tk.LabelFrame(self.win_hitung, text="Daftar Bahan", bg="white", padx=10, pady=10)
        frame_kanan.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        columns = ("Bahan", "Kategori", "Berat", "Kalori")
        self.tree = ttk.Treeview(frame_kanan, columns=columns, show="headings", height=12)
        
        self.tree.heading("Bahan", text="Bahan")
        self.tree.column("Bahan", width=120)
        self.tree.heading("Kategori", text="Kategori")
        self.tree.column("Kategori", width=80)
        self.tree.heading("Berat", text="Gr")
        self.tree.column("Berat", width=50, anchor="center")
        self.tree.heading("Kalori", text="Kcal")
        self.tree.column("Kalori", width=60, anchor="center")
        
        self.tree.pack(fill="both", expand=True)

        self.lbl_total = tk.Label(frame_kanan, text="Total: 0 Kcal", font=("Arial", 14, "bold"), bg="white", fg="red")
        self.lbl_total.pack(pady=10, anchor="e")
        tk.Button(frame_kanan, text="💾 Export Laporan Cepat", command=self.simpan_laporan_excel, bg="#4CAF50", fg="white").pack(fill="x", pady=5)

    def simpan_bahan_baru(self):
        nama = self.ent_nama_baru.get()
        kat = self.cb_kategori_baru.get()
        kal = self.ent_kalori_baru.get()

        if nama and kat and kal:
            try:
                self.db_bahan.append({
                    "nama": nama, "kategori": kat, "kalori": float(kal)
                })
                messagebox.showinfo("Sukses", "Bahan berhasil ditambahkan!")
                self.ent_nama_baru.delete(0, 'end')
                self.ent_kalori_baru.delete(0, 'end')
            except ValueError:
                messagebox.showerror("Error", "Kalori harus berupa angka")
        else:
            messagebox.showwarning("Kosong", "Isi semua data!")

    def tambah_ke_resep(self):
        nama_bahan = self.cb_pilih_bahan.get()
        if not nama_bahan: return
        data_ketemu = None
        for item in self.db_bahan:
            if item['nama'] == nama_bahan:
                data_ketemu = item
                break
        
        if data_ketemu:
            gram = int(self.var_gram.get())
            total_kal = (gram / 100) * data_ketemu['kalori']
            self.data_resep.append({
                "Bahan": nama_bahan,
                "Kategori": data_ketemu['kategori'],
                "Berat": gram,
                "Kalori": total_kal
            })
            self.tree.insert("", "end", values=(nama_bahan, data_ketemu['kategori'], gram, f"{total_kal:.1f}"))
            self.update_total_kalori()

    def update_total_kalori(self):
        # Hitung total bahan dasar
        total_bahan = 0
        for item in self.data_resep:
            total_bahan += item['Kalori']
  
        try:
            porsi = int(self.var_porsi.get())
        except (ValueError, tk.TclError):
            porsi = 1

        grand_total = total_bahan * porsi
        
        self.lbl_total.config(text=f"Total ({porsi} Porsi): {grand_total:.1f} Kcal")

    def tampilkan_grafik(self):
        if not self.data_resep:
            messagebox.showwarning("Kosong", "Belum ada resep!")
            return
        
        kategori_dict = {}
        for item in self.data_resep:
            kat = item['Kategori']
            cal = item['Kalori']
            if kat in kategori_dict:
                kategori_dict[kat] += cal
            else:
                kategori_dict[kat] = cal
        
        labels = list(kategori_dict.keys())
        values = list(kategori_dict.values())

        plt.figure(figsize=(10, 5))
        plt.suptitle("Analisis Gizi Resep (Perbandingan Komposisi)", fontsize=16)
    
        plt.subplot(1, 2, 1)
        plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=['#ff9999','#66b3ff','#99ff99','#ffcc99'])
        plt.title("Komposisi Kategori")

        plt.subplot(1, 2, 2)
        plt.bar(labels, values, color='#66b3ff')
        plt.title("Kalori per Kategori (Resep Dasar)")
        plt.ylabel("Kalori")

        plt.tight_layout()
        plt.show()

    def simpan_laporan_excel(self):
        if not self.data_resep: 
            messagebox.showwarning("Peringatan", "Data resep kosong!")
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path: return

        try:
            kategori_dict = {}
            for item in self.data_resep:
                kat = item['Kategori']
                cal = item['Kalori']
                kategori_dict[kat] = kategori_dict.get(kat, 0) + cal
            
            plt.figure(figsize=(8, 4))
            plt.bar(kategori_dict.keys(), kategori_dict.values(), color='skyblue')
            plt.title("Analisis Resep")
            plt.savefig("grafik_temp.png")
            plt.close()

            try:
                porsi = int(self.var_porsi.get())
            except:
                porsi = 1

            df = pd.DataFrame(self.data_resep)            
            
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            df.to_excel(writer, sheet_name='Laporan', startrow=5, index=False)
            
            ws = writer.book['Laporan']
            ws['A1'] = "NAMA RESEP:"
            ws['B1'] = self.var_nama_resep.get()
            ws['A2'] = "JUMLAH PORSI:"
            ws['B2'] = porsi
            ws['A3'] = "TANGGAL:"
            ws['B3'] = datetime.now().strftime("%Y-%m-%d")

            total_kal = sum(item['Kalori'] for item in self.data_resep) * porsi
            ws['A4'] = f"TOTAL KALORI ({porsi} PORSI): {total_kal:.1f} Kcal"
            
            from openpyxl.drawing.image import Image
            img = Image("grafik_temp.png")
            ws.add_image(img, 'F1')
            
            writer.close()

            if os.path.exists("grafik_temp.png"):
                os.remove("grafik_temp.png")
                
            messagebox.showinfo("Sukses", "Laporan berhasil diexport!")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def logout(self):
        answer = messagebox.askyesno("Konfirmasi", "Yakin mau logout?")
        if answer:
            self.window_menu.destroy()
            self.root.deiconify()

    def keluar_app(self):
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = AplikasiCaloriku(root)
    root.mainloop()